import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime
import sys
# A biblioteca openpyxl é necessária para escrever ficheiros .xlsx
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import Rule, DataBarRule, DifferentialStyle
from openpyxl.utils import get_column_letter
# --- NOVA ARQUITETURA: Importação do PyMuPDF (fitz) ---
import fitz  # PyMuPDF
from thefuzz import fuzz

# --- REATORAÇÃO: Constantes para nomes de status ---
STATUS_OPEN = "OPEN"
STATUS_CLOSED = "CLOSED"
STATUS_WAIT_APPROVAL = "WAIT APPROVAL"
STATUS_POSTPONED = "POSTPONED"
STATUS_REPLANEJADO = "REPLANEJADO"
STATUS_RETIRADA = "RETIRADA"


def extrair_dados_pdf_pymupdf(caminho_pdf):
    """
    Extrai dados de tabelas de um PDF usando a arquitetura robusta do PyMuPDF,
    projetada para lidar com tabelas que se estendem por várias páginas.
    """
    dados_cabecalho = {"report_date": None}
    try:
        doc = fitz.open(caminho_pdf)
        page_one = doc[0]
        text_page_one = page_one.get_text()

        # Extração de data (lógica mantida)
        report_date = None
        match1 = re.search(
            r"Today\s+([\w\s,]+\d{4})", text_page_one, re.IGNORECASE)
        if match1:
            date_str_raw = match1.group(1).strip()
            date_str_clean = date_str_raw.replace(',', '')
            for fmt in ["%B %d %Y", "%b %d %Y"]:
                try:
                    report_date = datetime.strptime(date_str_clean, fmt)
                    print(
                        f"INFO: Data encontrada (Padrão 1): '{date_str_raw}'")
                    break
                except ValueError:
                    continue
        if not report_date:
            match2 = re.search(
                r"Today\s+(\d{2}/\d{2}/\d{4})", text_page_one, re.IGNORECASE)
            if match2:
                date_str = match2.group(1).strip()
                try:
                    report_date = datetime.strptime(date_str, "%d/%m/%Y")
                    print(f"INFO: Data encontrada (Padrão 2): '{date_str}'")
                except ValueError:
                    pass

        dados_cabecalho["report_date"] = report_date or datetime.now()

    except Exception as e:
        print(
            f"Aviso: Não foi possível ler o cabeçalho do PDF. Erro: {e}. Usando data atual.")
        dados_cabecalho["report_date"] = datetime.now()

    # --- CORREÇÃO DEFINITIVA (SEQ 53): Arquitetura de Validação na Fonte ---
    # ETAPA 1: Extração Bruta e Filtro de Integridade Imediato
    validated_rows = []
    # Adicione outros grupos válidos se necessário
    VALID_GROUPS = {"Planned", "Internal Procedure", "Customer Request"}
    ALL_STATUSES = {STATUS_OPEN, STATUS_CLOSED, STATUS_WAIT_APPROVAL,
                    STATUS_POSTPONED, STATUS_REPLANEJADO, STATUS_RETIRADA}

    try:
        doc = fitz.open(caminho_pdf)
        for page_num in range(1, len(doc)):
            page = doc[page_num]
            tables_on_page = page.find_tables()
            if not tables_on_page:
                continue

            raw_table_data = tables_on_page[0].extract()
            header_signature = ['SEQ', 'GROUP', 'DESCRIPTION']

            for row in raw_table_data:
                if any(sig in str(cell) for sig, cell in zip(header_signature, row)):
                    continue

                # --- LÓGICA DE EXTRAÇÃO REVISADA E MAIS ROBUSTA ---
                id_val = str(row[0] or '').strip() if len(row) > 0 else ''
                seq_val_c = str(row[1] or '').strip() if len(row) > 1 else ''
                is_critical_issue = id_val.isdigit() and seq_val_c.isdigit()

                seq_val_n = str(row[1] or '').strip() if len(row) > 1 else ''
                is_task_normal = seq_val_n.isdigit()

                seq_val_s = str(row[0] or '').strip() if len(row) > 0 else ''
                is_task_shifted = seq_val_s.isdigit()

                if is_critical_issue:
                    description = str(row[2] or '').strip() if len(
                        row) > 2 else ''
                    status = str(row[3] or '').strip() if len(
                        row) > 3 else STATUS_OPEN

                    if status not in ALL_STATUSES and len(status) > 20:
                        description = (description + ' ' + status).strip()
                        status = STATUS_OPEN
                    elif status not in ALL_STATUSES:
                        status = STATUS_OPEN

                    normalized_row = [None, seq_val_c, 'Finding',
                                      description, status, id_val, None]
                    validated_rows.append(normalized_row)

                elif is_task_normal:
                    group_val = str(row[2] or '').strip() if len(
                        row) > 2 else ''
                    if group_val in VALID_GROUPS:
                        validated_rows.append(list(row))
                    else:
                        seq = seq_val_n
                        phase = row[0]
                        content_cells = row[2:]
                        full_text = ' '.join(str(c or '').strip()
                                             for c in content_cells if c).strip()
                        group, description, status, external_task = "Finding", full_text, STATUS_OPEN, None

                        if description.startswith("SB/ADs"):
                            group = "SB/ADs"
                            description = description.replace(
                                "SB/ADs", "", 1).strip()
                        elif description.startswith("Customer Report"):
                            group = "Customer Report"
                            description = description.replace(
                                "Customer Report", "", 1).strip()

                        temp_desc = description
                        for s in ALL_STATUSES:
                            if f" {s} " in f" {temp_desc} " or temp_desc.endswith(f" {s}"):
                                parts = temp_desc.rsplit(s, 1)
                                description = parts[0].strip()
                                status = s
                                potential_et = parts[1].strip()
                                if potential_et:
                                    external_task = potential_et
                                break

                        if not external_task and group == "SB/ADs":
                            match = re.search(
                                r"(AD\s?\(ANAC\)\s?\d{4}-\d{2}-\d{2})", description)
                            if match:
                                external_task = match.group(1).strip()

                        validated_rows.append(
                            [phase, seq, group, description, status, external_task, None])

                elif is_task_shifted and not is_critical_issue:
                    group_val = str(row[1] or '').strip() if len(
                        row) > 1 else ''
                    if group_val in VALID_GROUPS:
                        validated_rows.append([None] + list(row))
                    else:
                        seq = seq_val_s
                        content_cells = row[1:]
                        full_text = ' '.join(str(c or '').strip()
                                             for c in content_cells if c).strip()
                        group, description, status, external_task = "Finding", full_text, STATUS_OPEN, None

                        if description.startswith("SB/ADs"):
                            group = "SB/ADs"
                            description = description.replace(
                                "SB/ADs", "", 1).strip()
                        elif description.startswith("Customer Report"):
                            group = "Customer Report"
                            description = description.replace(
                                "Customer Report", "", 1).strip()

                        temp_desc = description
                        for s in ALL_STATUSES:
                            if f" {s} " in f" {temp_desc} " or temp_desc.endswith(f" {s}"):
                                parts = temp_desc.rsplit(s, 1)
                                description = parts[0].strip()
                                status = s
                                potential_et = parts[1].strip()
                                if potential_et:
                                    external_task = potential_et
                                break

                        validated_rows.append(
                            [None, seq, group, description, status, external_task, None])

                elif validated_rows:
                    continuation_text = ' '.join(str(c or '').replace(
                        '\n', ' ').strip() for c in row if c is not None and str(c).strip())
                    is_new_task_code = re.match(
                        r"^\d{2}-\d{2}-\d{2}-\d{3}", continuation_text.strip())
                    is_header_text = 'PHASE SEQ GROUP' in continuation_text

                    if is_new_task_code or is_header_text:
                        print(
                            f"AVISO: Linha ignorada para evitar corrupção da tarefa anterior. Conteúdo: '{continuation_text[:100]}...'")
                        continue

                    if continuation_text:
                        while len(validated_rows[-1]) <= 3:
                            validated_rows[-1].append('')
                        current_description = validated_rows[-1][3]
                        validated_rows[-1][3] = (str(current_description or '') +
                                                 ' ' + continuation_text).strip()
        doc.close()

    except Exception as e:
        print(f"Erro ao extrair tabelas com PyMuPDF: {e}")
        return dados_cabecalho, pd.DataFrame()

    if not validated_rows:
        return dados_cabecalho, pd.DataFrame()

    colunas = ['PHASE', 'SEQ', 'GROUP', 'DESCRIPTION',
               'STATUS', 'EXTERNAL TASK', 'ORIG']
    df_final = pd.DataFrame(validated_rows)

    num_cols_to_assign = min(len(df_final.columns), len(colunas))
    df_final = df_final.iloc[:, :num_cols_to_assign]
    df_final.columns = colunas[:num_cols_to_assign]

    if df_final.empty:
        return dados_cabecalho, df_final

    original_cols = [col for col in colunas if col in df_final.columns]

    for col in original_cols:
        if df_final[col].dtype == 'object':
            df_final[col] = df_final[col].astype(str).str.replace(
                r'\s+', ' ', regex=True).str.strip().replace('nan', '')

    df_final['SEQ'] = pd.to_numeric(df_final['SEQ'], errors='coerce')
    df_final.dropna(subset=['SEQ'], inplace=True)
    df_final['SEQ'] = df_final['SEQ'].astype(int)

    def prioritize_group(series):
        if 'Customer Report' in series.values:
            return 'Customer Report'
        if 'SB/ADs' in series.values:
            return 'SB/ADs'
        if 'Planned' in series.values:
            return 'Planned'
        return series.iloc[0]

    def prioritize_description(series):
        descriptions = pd.Series(series).str.strip().dropna().unique()
        descriptions = [d for d in descriptions if d]
        if not descriptions:
            return ""
        clean_descriptions = [
            d for d in descriptions if 'PHASE SEQ GROUP' not in d and 'Assunto escalado' not in d and 'MATERIAL SEM PRAZO' not in d]
        if clean_descriptions:
            return min(clean_descriptions, key=len)
        return max(descriptions, key=len)

    agg_dict = {'PHASE': 'first', 'GROUP': prioritize_group, 'DESCRIPTION': prioritize_description,
                'STATUS': 'first', 'EXTERNAL TASK': 'first', 'ORIG': 'first'}
    agg_dict_filtered = {k: v for k,
                         v in agg_dict.items() if k in df_final.columns}
    df_final = df_final.groupby('SEQ', as_index=False).agg(agg_dict_filtered)
    df_final = df_final.reindex(columns=original_cols)
    df_final.loc[df_final['STATUS'] == '', 'STATUS'] = STATUS_WAIT_APPROVAL
    return dados_cabecalho, df_final


if __name__ == "__main__":
    nome_pasta_relatorios = 'Relatorios_PDF'
    if not os.path.isdir(nome_pasta_relatorios):
        print(f"❌ ERRO: A pasta '{nome_pasta_relatorios}' não foi encontrada.")
        sys.exit()

    arquivos_candidatos = [os.path.join(nome_pasta_relatorios, f) for f in os.listdir(
        nome_pasta_relatorios) if f.lower().startswith('customer_report') and f.lower().endswith('.pdf')]
    if not arquivos_candidatos:
        print(
            f"❌ ERRO: Nenhum ficheiro PDF 'Customer_Report' encontrado na pasta '{nome_pasta_relatorios}'.")
        sys.exit()

    arquivos_ordenados = sorted(arquivos_candidatos, key=os.path.getmtime)
    print(
        f"📄 Encontrados {len(arquivos_ordenados)} relatórios para processar.")
    nome_arquivo_mestre = 'Dashboard_Mestre.xlsx'
    ids_antigos = set()

    if os.path.exists(nome_arquivo_mestre):
        print(
            f"📖 Verificando o ficheiro mestre existente: '{nome_arquivo_mestre}'")
        try:
            df_antigo = pd.read_excel(
                nome_arquivo_mestre, sheet_name='Dashboard', skiprows=11)
            if 'SEQ' in df_antigo.columns:
                df_antigo['SEQ'] = pd.to_numeric(
                    df_antigo['SEQ'], errors='coerce')
                df_antigo.dropna(subset=['SEQ'], inplace=True)
                df_antigo['SEQ'] = df_antigo['SEQ'].astype(int)
                ids_antigos = set(df_antigo['SEQ'].astype(str))
                print(
                    f"   -> {len(ids_antigos)} tarefas encontradas na versão anterior.")
        except Exception as e:
            print(
                f"   -> Aviso: Não foi possível ler o ficheiro mestre anterior: {e}.")

    df_mestre = pd.DataFrame()
    print("✨ Criando novo dashboard...")

    data_ultimo_relatorio = None

    for arquivo_pdf in arquivos_ordenados:
        print(f"\n--- Processando: '{os.path.basename(arquivo_pdf)}' ---")
        dados_cabecalho, df_novo = extrair_dados_pdf_pymupdf(arquivo_pdf)
        if df_novo.empty:
            print(
                f"⚠️ Nenhuma tarefa encontrada em '{os.path.basename(arquivo_pdf)}'.")
            continue

        data_relatorio = dados_cabecalho['report_date']
        data_ultimo_relatorio = data_relatorio
        df_novo['UniqueID'] = df_novo['SEQ'].astype(str)

        if not df_mestre.empty:
            ids_mestre = set(df_mestre['UniqueID'])
            ids_novo = set(df_novo['UniqueID'])
            ids_retirados = ids_mestre - ids_novo
            is_already_handled = df_mestre['STATUS'].isin(
                [STATUS_CLOSED, STATUS_RETIRADA])
            idx_retirados = df_mestre['UniqueID'].isin(
                ids_retirados) & ~is_already_handled
            if idx_retirados.any():
                df_mestre.loc[idx_retirados, 'STATUS'] = STATUS_RETIRADA
                df_mestre.loc[idx_retirados,
                              'Data Fechamento'] = data_relatorio

        for _, row_nova in df_novo.iterrows():
            unique_id = row_nova['UniqueID']
            if not df_mestre.empty and unique_id in df_mestre['UniqueID'].values:
                idx = df_mestre.index[df_mestre['UniqueID'] == unique_id][0]

                status_antigo = df_mestre.at[idx, 'STATUS']
                status_novo = row_nova['STATUS']

                df_mestre.at[idx, 'STATUS'] = status_novo
                df_mestre.at[idx, 'DESCRIPTION'] = row_nova['DESCRIPTION']
                df_mestre.at[idx, 'EXTERNAL TASK'] = row_nova['EXTERNAL TASK']
                df_mestre.at[idx, 'GROUP'] = row_nova['GROUP']

                data_fechamento_atual = df_mestre.at[idx, 'Data Fechamento']
                is_paused_novo = status_novo in [
                    STATUS_POSTPONED, STATUS_REPLANEJADO]
                was_not_paused_before = status_antigo not in [
                    STATUS_POSTPONED, STATUS_REPLANEJADO]

                if is_paused_novo and was_not_paused_before and pd.isna(data_fechamento_atual):
                    df_mestre.at[idx, 'Data Fechamento'] = data_relatorio
                elif status_novo == STATUS_CLOSED and pd.isna(data_fechamento_atual):
                    df_mestre.at[idx, 'Data Fechamento'] = data_relatorio
                elif status_novo not in [STATUS_CLOSED, STATUS_POSTPONED, STATUS_REPLANEJADO] and pd.notna(data_fechamento_atual):
                    df_mestre.at[idx, 'Data Fechamento'] = pd.NaT

            else:
                nova_linha = row_nova.to_dict()
                nova_linha['Data Abertura'] = data_relatorio
                nova_linha['Última Atualização'] = data_relatorio
                nova_linha['Data Fechamento'] = data_relatorio if nova_linha['STATUS'] == STATUS_CLOSED else pd.NaT
                df_mestre = pd.concat(
                    [df_mestre, pd.DataFrame([nova_linha])], ignore_index=True)

        if 'UniqueID' not in df_mestre.columns and not df_mestre.empty:
            df_mestre['UniqueID'] = df_mestre['SEQ'].astype(str)

    if not df_mestre.empty:
        if data_ultimo_relatorio is not None:
            df_mestre['Última Atualização'] = data_ultimo_relatorio

        idx_closed_no_date = (df_mestre['STATUS'] == STATUS_CLOSED) & (
            df_mestre['Data Fechamento'].isna())
        if idx_closed_no_date.any():
            df_mestre.loc[idx_closed_no_date,
                          'Data Fechamento'] = df_mestre.loc[idx_closed_no_date, 'Última Atualização']

        df_mestre['is_new'] = df_mestre['UniqueID'].apply(
            lambda x: x not in ids_antigos)

        hoje = datetime.now()
        df_mestre['Data Abertura_dt'] = pd.to_datetime(
            df_mestre['Data Abertura'], errors='coerce')
        df_mestre['Data Fechamento_dt'] = pd.to_datetime(
            df_mestre['Data Fechamento'], errors='coerce')
        df_mestre['Última Atualização_dt'] = pd.to_datetime(
            df_mestre['Última Atualização'], errors='coerce')

        idx_com_data_fim = df_mestre['Data Fechamento_dt'].notna()
        df_mestre.loc[idx_com_data_fim, 'Dias em Aberto'] = (
            df_mestre.loc[idx_com_data_fim, 'Data Fechamento_dt'] - df_mestre.loc[idx_com_data_fim, 'Data Abertura_dt']).dt.days

        idx_sem_data_fim = df_mestre['Data Fechamento_dt'].isna()
        df_mestre.loc[idx_sem_data_fim, 'Dias em Aberto'] = (
            hoje - df_mestre.loc[idx_sem_data_fim, 'Data Abertura_dt']).dt.days

        df_mestre['Dias em Aberto'] = pd.to_numeric(
            df_mestre['Dias em Aberto'], errors='coerce').astype('Int64')

        df_mestre.sort_values(by='SEQ', inplace=True)
        df_mestre.reset_index(drop=True, inplace=True)

        print("\n🔍 Analisando similaridade de 'DESCRIPTION'...")
        LIMITE_SIMILARIDADE = 98
        descricoes = df_mestre['DESCRIPTION'].dropna().astype(str).tolist()
        indices_para_colorir = set()
        for i in range(len(descricoes)):
            for j in range(i + 1, len(descricoes)):
                if fuzz.ratio(descricoes[i], descricoes[j]) >= LIMITE_SIMILARIDADE:
                    indices_para_colorir.add(i)
                    indices_para_colorir.add(j)
        if indices_para_colorir:
            print(
                f"   -> {len(indices_para_colorir)} tarefas com descrição similar encontradas.")
        else:
            print("   -> Nenhuma tarefa com descrição similar encontrada.")

        is_new_series = df_mestre['is_new']

        total_tarefas = len(df_mestre)
        count_fechadas = len(df_mestre[df_mestre['STATUS'] == STATUS_CLOSED])
        count_retiradas = len(
            df_mestre[df_mestre['STATUS'] == STATUS_RETIRADA])
        count_nao_aprov = len(
            df_mestre[df_mestre['STATUS'] == STATUS_WAIT_APPROVAL])
        count_replanejadas = len(
            df_mestre[df_mestre['STATUS'].isin([STATUS_POSTPONED, STATUS_REPLANEJADO])])
        count_abertas = len(df_mestre[df_mestre['STATUS'] == STATUS_OPEN])
        percentual_conclusao = (
            count_fechadas + count_retiradas) / total_tarefas if total_tarefas > 0 else 0

        colunas_finais = ['GROUP', 'SEQ', 'DESCRIPTION', 'STATUS', 'EXTERNAL TASK', 'ORIG',
                          'Data Abertura', 'Data Fechamento', 'Última Atualização', 'Dias em Aberto']
        df_mestre_excel = df_mestre[colunas_finais].copy()
        df_mestre_excel['Data Fechamento'] = df_mestre_excel['Data Fechamento'].astype(
            object)
        idx_replanejado = df_mestre_excel['STATUS'].isin(
            [STATUS_POSTPONED, STATUS_REPLANEJADO])
        # A exibição no Excel agora usará a 'Data Fechamento' que foi definida corretamente
        df_mestre_excel.loc[idx_replanejado, 'Data Fechamento'] = df_mestre.loc[idx_replanejado, 'Data Fechamento'].apply(
            lambda x: f"Replanejado em {x.strftime('%d/%m/%Y')}" if pd.notna(x) else "Replanejado")
        idx_retiradas = df_mestre_excel['STATUS'] == STATUS_RETIRADA
        df_mestre_excel.loc[idx_retiradas, 'Data Fechamento'] = df_mestre.loc[idx_retiradas, 'Data Fechamento'].apply(
            lambda x: f"Retirada em {x.strftime('%d/%m/%Y')}" if pd.notna(x) else "Retirada")

        try:
            with pd.ExcelWriter(nome_arquivo_mestre, engine='openpyxl') as writer:
                df_mestre_excel.to_excel(
                    writer, sheet_name='Dashboard', startrow=11, index=False)
                workbook = writer.book
                worksheet = writer.sheets['Dashboard']

                # --- CABEÇALHO ---
                font_titulo = Font(name='Roboto', size=16,
                                   bold=True, color='FFFFFF')
                fill_titulo = PatternFill(
                    start_color='215C98', end_color='215C98', fill_type='solid')
                align_center = Alignment(
                    horizontal='center', vertical='center')
                align_left = Alignment(horizontal='left', vertical='center')
                align_right = Alignment(horizontal='right', vertical='center')
                borda_fina = Border(left=Side(style='thin'), right=Side(
                    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                worksheet.merge_cells('A1:C1')
                cell_titulo = worksheet['A1']
                cell_titulo.value = 'Dashboard Mestre de Acompanhamento de Tarefas'
                cell_titulo.font = font_titulo
                cell_titulo.fill = fill_titulo
                cell_titulo.alignment = align_center

                worksheet.merge_cells('A2:C2')
                cell_timestamp = worksheet['A2']
                cell_timestamp.value = f"Última atualização: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                cell_timestamp.font = Font(
                    name='Roboto', size=9, italic=True, color='808080')
                cell_timestamp.alignment = align_center

                # --- SUMÁRIO COM ÍCONES COLORIDOS ---
                font_label_default = Font(name='Roboto', size=11, bold=True)
                font_fechadas = Font(name='Roboto', size=11,
                                     bold=True, color="00B050")
                font_retiradas = Font(
                    name='Roboto', size=11, bold=True, color="C00000")
                font_abertas = Font(name='Roboto', size=11,
                                    bold=True, color="0070C0")
                font_nao_aprov = Font(
                    name='Roboto', size=11, bold=True, color="9C6500")
                font_replanejadas = Font(
                    name='Roboto', size=11, bold=True, color="008B8B")

                summary_data = [
                    {'ref': 'A4', 'label': '📈 Progresso Geral:', 'value_ref': 'B4',
                     'value': percentual_conclusao, 'font': font_label_default},
                    {'ref': 'A5', 'label': '✅ Tarefas Fechadas:', 'value_ref': 'B5',
                     'value': count_fechadas, 'font': font_fechadas},
                    {'ref': 'A6', 'label': '❌ Tarefas Retiradas:', 'value_ref': 'B6',
                     'value': count_retiradas, 'font': font_retiradas},
                    {'ref': 'A7', 'label': '📋 Tarefas Abertas:', 'value_ref': 'B7',
                     'value': count_abertas, 'font': font_abertas},
                    {'ref': 'A8', 'label': '⏳ Não Aprovadas:', 'value_ref': 'B8',
                     'value': count_nao_aprov, 'font': font_nao_aprov},
                    {'ref': 'A9', 'label': '🔄 Tarefas Replanejadas:', 'value_ref': 'B9',
                     'value': count_replanejadas, 'font': font_replanejadas},
                    {'ref': 'A10', 'label': '🎯 Total de Tarefas:', 'value_ref': 'B10',
                     'value': total_tarefas, 'font': font_label_default},
                ]

                for item in summary_data:
                    label_cell = worksheet[item['ref']]
                    label_cell.value = item['label']
                    label_cell.font = item['font']
                    label_cell.alignment = align_left

                    value_cell = worksheet[item['value_ref']]
                    value_cell.value = item['value']
                    value_cell.font = font_label_default
                    value_cell.alignment = align_right

                worksheet['B4'].number_format = '0.00%'
                worksheet.conditional_formatting.add('B4', DataBarRule(
                    start_type='num', start_value=0, end_type='num', end_value=1, color="00B050", showValue=True))

                # --- LEGENDA DE CORES ---
                worksheet['D2'].value = "Legenda de Cores:"
                worksheet['D2'].font = Font(
                    name='Roboto', size=11, bold=True, underline="single")
                worksheet['D2'].alignment = align_left
                legend_data = [
                    ("Tarefa Retirada", "C00000"), ("Tarefa Fechada", "00B050"),
                    ("Aguardando Aprovação", "FFFF00"), ("Tarefa Aberta", "ADD8E6"),
                    (f"Tarefa {STATUS_REPLANEJADO.capitalize()}",
                     "00FFFF"), ("Nova Tarefa", "FA8072"),
                    ("Descrição Similar (>=98%)", "4B0082")
                ]
                for i, (label, color_hex) in enumerate(legend_data):
                    worksheet[f'D{3+i}'].value = label
                    worksheet[f'D{3+i}'].font = Font(name='Roboto', size=10)
                    worksheet[f'E{3+i}'].fill = PatternFill(
                        start_color=color_hex, end_color=color_hex, fill_type="solid")
                    worksheet[f'E{3+i}'].border = borda_fina

                # --- FORMATAÇÃO DA TABELA PRINCIPAL ---
                font_cabecalho_tabela = Font(
                    name='Roboto', bold=True, color='FFFFFF')
                fill_cabecalho_tabela = PatternFill(
                    start_color='215C98', end_color='215C98', fill_type='solid')

                for cell in worksheet["12:12"]:
                    cell.font = font_cabecalho_tabela
                    cell.fill = fill_cabecalho_tabela
                    cell.alignment = align_center
                    cell.border = borda_fina

                table_range = f"A12:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
                worksheet.auto_filter.ref = table_range

                max_row = worksheet.max_row
                for row in worksheet.iter_rows(min_row=13, max_row=max_row):
                    for cell in row:
                        cell.border = borda_fina
                        cell.alignment = Alignment(
                            vertical='center', wrap_text=True, horizontal='center')

                for col_letter in ['G', 'I', 'H']:
                    for cell in worksheet[col_letter]:
                        if cell.row > 12 and not isinstance(cell.value, str):
                            cell.number_format = 'dd/mm/yyyy'

                dxf_retirada = DifferentialStyle(
                    font=Font(color="FFFFFF"), fill=PatternFill(bgColor="C00000"))
                dxf_closed = DifferentialStyle(
                    font=Font(color="FFFFFF"), fill=PatternFill(bgColor="00B050"))
                dxf_wait = DifferentialStyle(
                    fill=PatternFill(bgColor="FFFF00"))
                dxf_open = DifferentialStyle(
                    fill=PatternFill(bgColor="ADD8E6"))
                dxf_postponed = DifferentialStyle(
                    fill=PatternFill(bgColor="00FFFF"))

                range_total = f"A13:J{worksheet.max_row}"
                worksheet.conditional_formatting.add(range_total, Rule(type="expression", formula=[
                    f'$D13="{STATUS_RETIRADA}"'], stopIfTrue=True, dxf=dxf_retirada))
                worksheet.conditional_formatting.add(range_total, Rule(type="expression", formula=[
                    f'OR($D13="{STATUS_POSTPONED}", $D13="{STATUS_REPLANEJADO}")'], stopIfTrue=True, dxf=dxf_postponed))

                range_status = f"D13:D{worksheet.max_row}"
                worksheet.conditional_formatting.add(range_status, Rule(
                    type="cellIs", operator="equal", formula=[f'"{STATUS_OPEN}"'], dxf=dxf_open))
                worksheet.conditional_formatting.add(range_status, Rule(
                    type="cellIs", operator="equal", formula=[f'"{STATUS_CLOSED}"'], dxf=dxf_closed))
                worksheet.conditional_formatting.add(range_status, Rule(
                    type="cellIs", operator="equal", formula=[f'"{STATUS_WAIT_APPROVAL}"'], dxf=dxf_wait))
                worksheet.conditional_formatting.add(f"J13:J{worksheet.max_row}", Rule(
                    type="expression", formula=['AND($G13<>"", $H13<>"")'], dxf=dxf_closed))

                fill_new_seq = PatternFill(
                    start_color="FA8072", end_color="FA8072", fill_type="solid")
                for index, is_new in is_new_series.items():
                    if is_new:
                        excel_row = index + 13
                        worksheet[f'B{excel_row}'].fill = fill_new_seq
                        worksheet[f'C{excel_row}'].fill = fill_new_seq

                if indices_para_colorir:
                    fill_similar = PatternFill(
                        start_color="4B0082", end_color="4B0082", fill_type="solid")
                    font_similar = Font(color="FFFFFF", name='Roboto')
                    for index_df in indices_para_colorir:
                        cell = worksheet[f'C{index_df + 13}']
                        cell.fill = fill_similar
                        cell.font = font_similar

                worksheet.column_dimensions['A'].width = 25.5
                worksheet.column_dimensions['B'].width = 15
                worksheet.column_dimensions['C'].width = 60
                worksheet.column_dimensions['D'].width = 25
                worksheet.column_dimensions['E'].width = 30
                worksheet.column_dimensions['F'].width = 20
                worksheet.column_dimensions['G'].width = 20
                worksheet.column_dimensions['H'].width = 25
                worksheet.column_dimensions['I'].width = 20
                worksheet.column_dimensions['J'].width = 15
                worksheet.sheet_view.zoomScale = 70
                worksheet.freeze_panes = 'A13'

            print(
                f"\n✅ Dashboard mestre salvo e atualizado com sucesso em: '{nome_arquivo_mestre}'")
        except Exception as e:
            print(f"\n❌ ERRO ao salvar o ficheiro Excel: {e}")
            print("Verifique se o ficheiro 'Dashboard_Mestre.xlsx' não está aberto.")
    else:
        print("\n❌ Nenhuma tarefa foi extraída. O ficheiro mestre não foi alterado.")
